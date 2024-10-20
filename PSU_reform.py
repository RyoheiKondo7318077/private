import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart

para = 38

# パスで指定したファイルの一覧をリスト形式で取得. （ここでは一階層下のtestファイル以下）
csv_files = glob.glob('tsurumi/*.csv')

# 時間順でソート
csv_files.sort(key=os.path.getmtime, reverse=False)

#読み込むファイルのリストを表示
for a in csv_files:
    print(a)

#csvファイルの中身を追加していくリストを用意
data_list = []

#読み込むファイルのリストを走査
for file in csv_files:
    # CSVファイルを読み込み、PSU_in_v列だけ取得
    filtered_data = pd.read_csv(file, usecols=["PSU_in_v"])
    
    # PSU_in_v列の値が38以上のデータだけを抽出してリストに追加
    filtered_data = filtered_data[filtered_data['PSU_in_v'] >= 38]

    # 欠損値（NaN）が含まれる行を削除して上詰め
    filtered_data = filtered_data.dropna().reset_index(drop=True)

    # フィルタリングされたデータをリストに追加
    data_list.append(filtered_data)

#リストを全て行方向に結合
#axis=0:行方向に結合, sort
df = pd.concat(data_list, axis=1)

# カラムをリネーム
df.columns =["EL1-1","EL1-2","EL1-3","EL1-4","EL2-1","EL2-2","EL2-3","EL2-4"]

df.to_excel("PSU_in_v_over"+str(para)+"V.xlsx")

"""
#ここからグラフ作成
file_path = "PSU_in_v_over"+str(para)+"V.xlsx"

## .xlsxの場合
wb = load_workbook(file_path)

## indexでシートを指定
ws = wb["Sheet1"]

# 折れ線グラフ
chart = LineChart()

# データ範囲の指定
# B列からI列までのデータ（カラム付き時系列データ）
data = Reference(ws, min_col=2, max_col=9, min_row=1, max_row=ws.max_row)
# インデックス（X軸）の範囲 A列
categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

# データとインデックス（カテゴリ）をグラフに追加
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# data titles
chart.x_axis.title = '総運転時間 [h]'
chart.y_axis.title = '電圧 [V]'

# 値の範囲を指定
chart.y_axis.scaling.min = 38
chart.y_axis.scaling.max = 46
chart.x_axis.scaling.min = 0
chart.x_axis.scaling.max = 600

# このセルに貼り付け
ws.add_chart(chart, 'I2')

wb.save('Graph.xlsx')
"""
